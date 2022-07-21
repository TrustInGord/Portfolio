#This is a Python script that first takes a value from an Excel workbook and posts it to twitter.

from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep

#Opens Excel Document for data reading
#TimeHarvest.xlsm must be changed to xlsm of choice.

#Specify Username and Password
USERNAME = ""
PASSWORD = ""

wb = load_workbook('./TimeHarvest.xlsm')
mySheet = wb['Dashboard']

#Takes value from closed Excel document
#Change cell or range as needed.
timeSaved = round((mySheet.cell(4, 9).value), 2)
print(timeSaved)
wb.close()

#Records value to a txt file.
outfile = open("./CollectionPlate.txt", 'w')
myPhrase = ('Approximately ', str(timeSaved), ' seconds have been saved. Without The Fluff')
print (myPhrase)
outfile.write(str(myPhrase))

outfile.close()


#Opens Twitter in Selenium.
#Requires Chromedriver, or the driver of your chosen web browser.
driver = webdriver.Chrome(executable_path="/chromedriver.exe")

#Opens the website
driver.get("https://twitter.com/login")
sleep(5)
driver.find_element_by_name('session[username_or_email]').send_keys(USERNAME)
driver.find_element_by_name('session[password]').send_keys(PASSWORD)
driver.find_element_by_name('session[password]').send_keys(Keys.RETURN)
sleep(3)


#Reads colection plate file for the seconds saved
f = open("/CollectionPlate.txt", 'r')

for word in f:
    if word == "\n":
        continue
    driver.find_element_by_xpath("//a[@data-testid='SideNav_NewTweet_Button']").click()
    sleep(1)
    driver.find_element_by_class_name("notranslate").click()
    driver.find_element_by_class_name("notranslate").send_keys(myPhrase)
    driver.find_element_by_xpath("//div[@data-testid='tweetButton']").click()
    sleep(5)

f.close()

driver.close()
