from openpyxl import Workbook, load_workbook

wb = load_workbook('C:/Users/Gordon/Desktop/Wonderful Moments/0 - EveryProject/TimeHarvest.xlsm')
mySheet = wb['Dashboard']

print (mySheet)
timeSaved = (mySheet.cell(4, 9).value)
print(timeSaved)



wb.close()

x=int(5)
